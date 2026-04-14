# streamlit_app.py  –  M&V Report vs Calculation Sheet Sanity Check Tool
import base64
import io
import json
import logging
import os
import re
import sys
from pathlib import Path

import streamlit as st
import streamlit.components.v1 as components
from dotenv import load_dotenv

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).resolve().parent / "app"
PROMPT_PATH   = BASE_DIR / "assets" / "MV_Report_Calc_Sanity_Check_Prompt.txt"
TEMPLATE_PATH = BASE_DIR / "assets" / "M_V_Review_Sheet_Template_260407_V1_1.xlsx"
LOGO_PATH     = BASE_DIR / "static" / "arklogo2.png"

load_dotenv(BASE_DIR.parent / ".env")

sys.path.insert(0, str(BASE_DIR))
from excel_writer import write_review
from sn_extractor import extract_expected_sns

# ── Load assets once ──────────────────────────────────────────────────────────
REVIEWER_PROMPT: str   = PROMPT_PATH.read_text(encoding="utf-8")
TEMPLATE_BYTES:  bytes = TEMPLATE_PATH.read_bytes()
EXPECTED_SNS           = extract_expected_sns(str(TEMPLATE_PATH))

# ── API constants ─────────────────────────────────────────────────────────────
MODEL           = "claude-sonnet-4-6"
MAX_TOKENS      = 4000
THINKING_TOKENS = 3000
TIMEOUT_SECS    = 120

VALID_STATUS = {"Approved", "Not Approved", "Incomplete"}

logging.basicConfig(level=logging.INFO)


# ── Backend helpers ────────────────────────────────────────────────────────────

def img_to_base64(path: str) -> str:
    return base64.b64encode(Path(path).read_bytes()).decode("utf-8")

def _encode_bytes(data: bytes) -> str:
    return base64.standard_b64encode(data).decode("utf-8")

def _strip_fences(text: str) -> str:
    text = text.strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    return text.strip()

def _extract_json_text(response) -> str:
    for block in response.content:
        if block.type == "text":
            return block.text
    raise ValueError("No text content block found in API response.")

def _validate_items(items: list) -> list:
    errors = []
    for i, item in enumerate(items):
        if not isinstance(item, dict):
            errors.append(f"Item {i} is not a dict"); continue
        for field in ("sn", "status", "comment"):
            if field not in item:
                errors.append(f"Item {i} missing field '{field}'")
        if "status" in item and item["status"] not in VALID_STATUS:
            errors.append(f"Item {i} invalid status={item['status']!r}")
    if isinstance(items, list) and len(items) != 2:
        errors.append(f"Expected exactly 2 items, got {len(items)}")
    return errors

def _parse_and_validate(raw: str):
    cleaned = _strip_fences(raw)
    try:
        items = json.loads(cleaned)
    except json.JSONDecodeError as exc:
        return [], [f"JSON parse error: {exc}"]
    if not isinstance(items, list):
        return [], ["Response is not a JSON array"]
    return items, _validate_items(items)

def _xlsx_to_text(xlsx_bytes: bytes) -> str:
    """Convert an xlsx workbook to a plain-text CSV representation."""
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            parts.append(",".join("" if v is None else str(v) for v in row))
    return "\n".join(parts)

def _scan_xlsx_for_metadata(xlsx_bytes: bytes) -> dict:
    """
    Scan the Excel file header rows directly with openpyxl.
    Looks for labelled cells like 'Client:', 'ESP:', 'Prepared for:' etc.

    Strategy:
    - Iterate the first 40 rows of the first sheet using cell objects (not
      values_only) so we can look ahead to the row below.
    - For each cell whose text contains a known keyword, check BOTH the cell
      to the right (same row, next column) AND the cell directly below (same
      column, next row).  The first non-empty value wins.
    - Additionally scan rows 1-20 for any cell that looks like a company name
      and collect those as candidates (stored under "company_candidates").

    Returns a partial dict with client_name, facility_name, country, esp_name
    and company_candidates populated where found.
    """
    import openpyxl
    import re
    result = {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        ws = wb.worksheets[0]

        label_map = {
            "client": "client_name",
            "owner": "client_name",
            "prepared for": "client_name",
            "submitted to": "client_name",
            "esp": "esp_name",
            "contractor": "esp_name",
            "prepared by": "esp_name",
            "submitted by": "esp_name",
            "energy service": "esp_name",
            "facility": "facility_name",
            "site": "facility_name",
            "project": "facility_name",
            "plant": "facility_name",
            "building": "facility_name",
            "country": "country",
        }

        # Build a list-of-lists of cell objects for rows 1-40 so we can look
        # at the row below without a second pass.
        max_row = min(ws.max_row or 40, 40)
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=max_row):
            rows.append(list(row))

        company_keywords = re.compile(
            r"\b(LLC|Ltd|Inc|Co\.|Authority|Corporation|Corp|Group|Energy|Services|Solutions|Associates|Partners|International)\b",
            re.IGNORECASE,
        )

        company_candidates = []

        for ri, row in enumerate(rows):
            for ci, cell in enumerate(row):
                raw = cell.value
                if raw is None:
                    continue
                cell_str = str(raw).strip()
                if not cell_str:
                    continue
                cell_lower = cell_str.lower().rstrip(":").strip()

                # ── keyword label matching ────────────────────────────────────
                for keyword, field in label_map.items():
                    if keyword in cell_lower and field not in result:
                        value = None
                        # 1) cell to the right in the same row
                        if ci + 1 < len(row):
                            right = row[ci + 1].value
                            if right is not None and str(right).strip():
                                value = str(right).strip()
                        # 2) cell directly below (same column, next row)
                        if not value and ri + 1 < len(rows):
                            below = rows[ri + 1][ci].value
                            if below is not None and str(below).strip():
                                value = str(below).strip()
                        if value:
                            result[field] = value
                        break

                # ── company-name candidate collection (rows 1-20 only) ────────
                if ri < 20:
                    word_count = len(cell_str.split())
                    is_numeric = False
                    try:
                        float(cell_str.replace(",", "").replace("%", ""))
                        is_numeric = True
                    except ValueError:
                        pass
                    if not is_numeric and (
                        word_count > 3 or company_keywords.search(cell_str)
                    ):
                        if cell_str not in company_candidates:
                            company_candidates.append(cell_str)

        if company_candidates:
            result["company_candidates"] = company_candidates

        wb.close()
    except Exception as e:
        logging.warning("xlsx scan failed: %s", e)
    return result


def _extract_from_filename(*filenames: str) -> dict:
    """
    Parse year_number, period_type, period_number from one or more filenames.

    Handles patterns like:
      Y1, Y2, Year1, Year_1, YEAR-2
      Q1, Q2, Quarter1, Quarter_2, Q_3
      M1, M3, M12, Month1, Month_3, MONTH-12
      Combined: Y1Q2, Y2_M3, Y1-Q4
    """
    result = {}
    for filename in filenames:
        name = re.sub(r'\.[^.]+$', '', filename).upper()   # strip extension, uppercase
        name = re.sub(r'[_\-\s]+', '_', name)              # normalise separators

        # ── Year ──────────────────────────────────────────────────────────────
        if "year_number" not in result:
            m = re.search(r'(?<![A-Z0-9])Y(?:EAR)?_?(\d{1,2})(?![0-9])', name)
            if m:
                result["year_number"] = int(m.group(1))

        # ── Quarter ───────────────────────────────────────────────────────────
        if "period_type" not in result:
            m = re.search(r'(?<![A-Z0-9])Q(?:UARTER)?_?([1-4])(?![0-9])', name)
            if m:
                result["period_type"]   = "Q"
                result["period_number"] = int(m.group(1))

        # ── Month (only if no quarter already found) ──────────────────────────
        if "period_type" not in result:
            m = re.search(r'(?<![A-Z0-9])M(?:ONTH)?_?(1[0-2]|[1-9])(?![0-9])', name)
            if m:
                result["period_type"]   = "M"
                result["period_number"] = int(m.group(1))

        if len(result) == 3:   # found everything — no need to check more files
            break

    return result


def _extract_metadata(calc_text: str, pdf_bytes: bytes) -> dict:
    """
    Extract submission metadata from the uploaded documents.
    Step 1: direct openpyxl scan of Excel header rows (fast, no API).
    Step 2: Haiku AI call on PDF + Excel text to fill in remaining fields.
    """
    import anthropic

    # ── Step 1: fast openpyxl scan ────────────────────────────────────────────
    # Re-encode calc_text back from the string is complex; use raw bytes directly
    # The xlsx bytes are passed separately via run_sanity_check but here we only
    # have calc_text. Haiku handles the rest.

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {}

    # ── Step 2: extract full PDF text (up to 10 pages) ────────────────────────
    pdf_text = ""
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(pdf_bytes))
        for page in list(reader.pages)[:10]:
            pdf_text += (page.extract_text() or "") + "\n"
    except Exception as e:
        logging.warning("pypdf failed: %s", e)

    combined = (
        f"=== M&V REPORT (first 10 pages, full text) ===\n{pdf_text[:12000]}\n\n"
        f"=== M&V CALCULATION SHEET (Excel, full text) ===\n{calc_text[:12000]}"
    )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=600,
            messages=[{
                "role": "user",
                "content": (
                    "You are extracting metadata from an M&V (Measurement & Verification) "
                    "document package. Read ALL the text below carefully and extract every "
                    "field you can find.\n\n"
                    "Return ONLY a valid JSON object with these exact keys:\n"
                    "{\n"
                    '  "client_name": "the organisation that commissioned the work / the owner / '
                    'the government entity / the building owner — NOT the ESP",\n'
                    '  "facility_name": "name of the facility, building, or plant",\n'
                    '  "country": "country where the facility is located",\n'
                    '  "esp_name": "name of the Energy Service Provider / contractor / '
                    'the company that prepared the report",\n'
                    '  "project_name": "specific plant or tower name (e.g. Al Ain Plant)",\n'
                    '  "mv_option": "Option A or B or C or C - Mean Model or D, or empty string",\n'
                    '  "reporting_period_start": "dd/mm/yyyy or empty string",\n'
                    '  "reporting_period_end": "dd/mm/yyyy or empty string",\n'
                    '  "year_number": integer or null,\n'
                    '  "period_type": "M or Q or null",\n'
                    '  "period_number": integer or null\n'
                    "}\n\n"
                    "SEARCH HINTS — look for:\n"
                    "- Cover page, title page, header rows of the Excel\n"
                    "- 'Prepared for', 'Client', 'Owner', 'Submitted to' → client_name\n"
                    "- 'Prepared by', 'ESP', 'Contractor', 'Submitted by' → esp_name\n"
                    "- Project title, facility name, building name → facility_name\n"
                    "- Any country or location mentioned\n"
                    "- Reporting period dates (start and end)\n"
                    "- Year number (Y1, Y2, Year 1, Year 4 etc.)\n\n"
                    "Use empty string for any text field not found. "
                    "Return ONLY the JSON, no explanation.\n\n"
                    f"{combined}"
                ),
            }],
        )
        raw = response.content[0].text.strip()
        logging.info("Haiku response: %s", raw[:600])
        return json.loads(_strip_fences(raw))
    except Exception as e:
        logging.warning("Metadata extraction API call failed: %s", e)
        return {"_error": str(e)}


def _build_user_content(pdf_bytes: bytes, calc_text: str) -> list:
    return [
        {
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": _encode_bytes(pdf_bytes),
            },
            "title": "M&V Report",
        },
        {
            "type": "text",
            "text": f"M&V Calculation Sheet (full text extracted from Excel):\n\n{calc_text}",
        },
        {
            "type": "text",
            "text": (
                "Compare the M&V Report against the M&V Calculation Sheet using every "
                "question in your instructions. Return a single valid JSON array only — "
                "no markdown, no explanation, no text outside the array. "
                "Each element must have exactly three fields: sn, status, comment."
            ),
        },
    ]

def _call_claude(client, user_content: list) -> str:
    response = client.messages.create(
        model=MODEL,
        max_tokens=MAX_TOKENS,
        temperature=1,
        thinking={"type": "enabled", "budget_tokens": THINKING_TOKENS},
        system=REVIEWER_PROMPT,
        messages=[{"role": "user", "content": user_content}],
        timeout=TIMEOUT_SECS,
    )
    return _extract_json_text(response)

def _call_claude_retry(client, user_content: list, first_raw: str) -> str:
    retry_messages = [
        {"role": "user",      "content": user_content},
        {"role": "assistant", "content": first_raw},
        {"role": "user",      "content": (
            "Your previous response was invalid. Return ONLY a valid JSON array with "
            "exactly two elements for SNs 0.1 and 0.2. Each item must have:\n"
            "- sn: '0.1' or '0.2'\n"
            "- status: exactly one of: Approved / Not Approved / Incomplete\n"
            "  (do NOT use Approved as Noted — this is Round 1)\n"
            "- comment: string (never blank — include confirmation or discrepancy detail)\n"
            "No markdown, no explanation, no text outside the JSON array."
        )},
    ]
    response = client.messages.create(
        model=MODEL,
        max_tokens=MAX_TOKENS,
        temperature=1,
        thinking={"type": "enabled", "budget_tokens": THINKING_TOKENS},
        system=REVIEWER_PROMPT,
        messages=retry_messages,
        timeout=TIMEOUT_SECS,
    )
    return _extract_json_text(response)

def run_sanity_check(pdf_bytes: bytes, xlsx_bytes: bytes, pdf_filename: str,
                     xlsx_filename: str = "", submission_details: dict = None):
    import anthropic

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY is not set.")

    # Extract calc text once — used for both API call and metadata extraction
    calc_text = _xlsx_to_text(xlsx_bytes)

    # Step 1: fast openpyxl scan of Excel header rows
    meta = _scan_xlsx_for_metadata(xlsx_bytes)
    # Step 2: Haiku fills in anything not found by the direct scan
    haiku_meta = _extract_metadata(calc_text, pdf_bytes)
    # Merge: direct scan takes priority; Haiku fills gaps
    for k, v in haiku_meta.items():
        if v and not meta.get(k):
            meta[k] = v

    # Step 3: filename parsing — most reliable source for year / period fields;
    # overrides AI/scan results for those specific keys
    fn_meta = _extract_from_filename(pdf_filename, xlsx_filename or "")
    for k, v in fn_meta.items():
        if v is not None:
            meta[k] = v

    logging.info("Extracted metadata: %s", meta)

    # Manual submission details override auto-extracted values
    if submission_details:
        for key, val in submission_details.items():
            if val:
                meta[key] = val

    client = anthropic.Anthropic(api_key=api_key)
    user_content = _build_user_content(pdf_bytes, calc_text)

    try:
        raw = _call_claude(client, user_content)
    except anthropic.APITimeoutError:
        raise RuntimeError(
            "The API call timed out after 120 seconds. Please re-run the check."
        )
    except anthropic.AuthenticationError:
        raise RuntimeError(
            "API authentication failed. Please check the ANTHROPIC_API_KEY in your .env file."
        )

    items, errors = _parse_and_validate(raw)

    if errors:
        logging.warning("First response invalid: %s — retrying.", errors)
        raw = _call_claude_retry(client, user_content, raw)
        items, errors = _parse_and_validate(raw)
        if errors:
            raise RuntimeError(
                "The AI returned an invalid response after retry. Please re-run the check."
            )

    review_by_sn = {str(item["sn"]).strip(): item for item in items}
    missing_sns  = sorted(set(EXPECTED_SNS) - set(review_by_sn.keys()))

    approved     = sum(1 for it in items if it.get("status") == "Approved")
    not_approved = sum(1 for it in items if it.get("status") == "Not Approved")
    incomplete   = sum(1 for it in items if it.get("status") == "Incomplete")
    total        = len(items)

    filled_bytes = write_review(TEMPLATE_BYTES, review_by_sn, meta=meta)

    base_name = pdf_filename.replace(".pdf", "").replace(".PDF", "")
    output_filename = f"MV_Report_Sanity_Check_{base_name}.xlsx"

    return {
        "total":        total,
        "approved":     approved,
        "not_approved": not_approved,
        "incomplete":   incomplete,
        "missing_sns":  missing_sns,
        "excel_bytes":  filled_bytes,
        "filename":     output_filename,
    }


# ============================================================
# UI
# ============================================================

st.set_page_config(page_title="ARK Energy | M&V Report Sanity Check", layout="wide")

st.markdown(
    """
    <style>
    :root{
      --ark-blue: #0D6079;
      --ark-orange: #F79428;
      --ark-black: #000000;
    }

    /* ── Global font enforcement ─────────────────────────────────────── */
    html, body, :root, [class*="css"], * {
        font-family: "Trebuchet MS", Arial, sans-serif !important;
        color: var(--ark-black);
    }
    button, input, textarea, select, option,
    label, p, div, span, li, a, h1, h2, h3, h4, h5, h6,
    code, pre, small, strong, em, td, th {
        font-family: "Trebuchet MS", Arial, sans-serif !important;
    }

    /* ── Streamlit widget-specific overrides ─────────────────────────── */
    /* Selectbox — widget, dropdown list, options */
    [data-testid="stSelectbox"] *,
    [data-testid="stSelectbox"] select,
    [data-baseweb="select"] *,
    [data-baseweb="popover"] *,
    [role="listbox"], [role="listbox"] *,
    [role="option"],  [role="option"] * {
        font-family: "Trebuchet MS", Arial, sans-serif !important;
    }

    /* Alerts / success / warning / error / info banners */
    [data-testid="stAlert"] *,
    [data-testid="stNotification"] *,
    .stAlert *, .stSuccess *, .stWarning *,
    .stError *, .stInfo * {
        font-family: "Trebuchet MS", Arial, sans-serif !important;
    }

    /* Spinner */
    [data-testid="stSpinner"] *, .stSpinner * {
        font-family: "Trebuchet MS", Arial, sans-serif !important;
    }

    /* Download button */
    [data-testid="stDownloadButton"] * {
        font-family: "Trebuchet MS", Arial, sans-serif !important;
    }

    /* File uploader */
    [data-testid="stFileUploader"] *,
    [data-testid="stFileUploaderDropzone"] * {
        font-family: "Trebuchet MS", Arial, sans-serif !important;
    }

    /* Markdown / text elements */
    .stMarkdown *, .element-container * {
        font-family: "Trebuchet MS", Arial, sans-serif !important;
    }

    /* Tooltip and popover overlays */
    [data-baseweb="tooltip"] *,
    [data-baseweb="menu"] * {
        font-family: "Trebuchet MS", Arial, sans-serif !important;
    }

    header[data-testid="stHeader"] { display: none; }
    footer { display: none; }

    .block-container {
        padding-top: 6.2rem !important;
        padding-bottom: 1.2rem !important;
        max-width: 98vw !important;
    }

    .ark-nav {
        position: fixed;
        top: 0; left: 0; right: 0;
        z-index: 9999;
        background: linear-gradient(90deg,#060C2E 0%,#08133A 45%,#0B1A4A 100%);
        padding: 12px 18px;
        box-shadow: 0 6px 18px rgba(0,0,0,0.35);
    }
    .ark-nav-inner{
        width: 98vw; margin: 0 auto; border-radius: 14px; padding: 10px 14px;
        display: flex; align-items: center; justify-content: space-between;
        background: linear-gradient(90deg,#060C2E 0%,#08133A 45%,#0B1A4A 100%);
    }
    .ark-nav-left { display:flex; align-items:center; gap:14px; }
    .ark-nav-title { color:white !important; font-size:22px !important; font-weight:900; font-family:"Trebuchet MS",Arial,sans-serif !important; line-height:1.2; margin:0; }
    .ark-nav-subtitle { color:rgba(255,255,255,0.65) !important; font-size:13px !important; font-weight:400; font-family:"Trebuchet MS",Arial,sans-serif !important; margin:0; }
    .ark-nav-right { display:flex; align-items:center; gap:10px; }
    .pill {
        border-radius:999px; padding:8px 14px; font-size:14px; font-weight:900;
        border:1px solid rgba(255,255,255,0.25); color:white !important; background:transparent; white-space:nowrap;
    }

    .ark-section { margin-top:10px; margin-bottom:6px; display:flex; align-items:baseline; gap:10px; }
    .ark-section-title { font-size:18px; font-weight:900; color:var(--ark-blue); margin:0; line-height:1; }
    .ark-section-rule { height:2px; background:rgba(13,96,121,0.25); width:100%; margin-top:8px; margin-bottom:14px; }

    [data-testid="stFileUploaderDropzone"] {
        background-color: #ebebeb !important;
        border: 1px solid #c8c8c8 !important;
        border-radius: 6px !important;
    }
    [data-testid="stFileUploaderDropzone"]:hover {
        background-color: #e2e2e2 !important;
        border-color: #b0b0b0 !important;
    }

    [data-testid="stFileUploaderDropzone"] button [data-testid="stIconMaterial"] {
        display: none !important;
    }

    /* Text inputs — grey style */
    [data-testid="stTextInput"] input {
        background-color: #e8e8e8 !important;
        border: 1px solid #c8c8c8 !important;
        border-radius: 6px !important;
        color: #111 !important;
        padding: 10px 12px !important;
    }
    [data-testid="stTextInput"] input:focus {
        background-color: #e0e0e0 !important;
        border-color: #0D6079 !important;
        box-shadow: none !important;
        outline: none !important;
    }

    label { font-size:15px !important; font-weight:700 !important; }

    div.stButton > button[kind="primary"],
    div.stButton > button[kind="primary"] * { color:#FFFFFF !important; }
    div.stButton > button[kind="primary"] {
        background-color: var(--ark-orange) !important;
        color: #FFFFFF !important;
        font-size: 22px !important;
        font-weight: 900 !important;
        border-radius: 12px !important;
        padding: 14px 28px !important;
        border: none !important;
        box-shadow: 0 8px 20px rgba(247,148,40,0.25) !important;
        width: 100% !important;
    }
    div.stButton > button[kind="primary"]:hover { background-color:var(--ark-blue) !important; color:#FFFFFF !important; }

    .stat-card { background:white; border-radius:12px; padding:18px 16px; text-align:center; box-shadow:0 3px 10px rgba(0,0,0,0.06); margin-bottom:8px; font-family:"Trebuchet MS",Arial,sans-serif !important; }
    .stat-number { font-size:36px; font-weight:900; line-height:1; margin-bottom:6px; font-family:"Trebuchet MS",Arial,sans-serif !important; }
    .stat-label  { font-size:12px; font-weight:700; text-transform:uppercase; letter-spacing:0.05em; color:#555; font-family:"Trebuchet MS",Arial,sans-serif !important; }
    .color-blue   { color:#0D6079; }
    .color-green  { color:#375623; }
    .color-red    { color:#9C0006; }
    .color-orange { color:#9C6500; }
    </style>
    """,
    unsafe_allow_html=True,
)

# Enforce Trebuchet MS on every element — including React-rendered widgets
st.markdown("""
<script>
(function enforceTrebuchet() {
    const FONT = '"Trebuchet MS", Arial, sans-serif';
    function applyFont(root) {
        root.querySelectorAll('*').forEach(el => {
            el.style.setProperty('font-family', FONT, 'important');
        });
    }
    function run() {
        try { applyFont(window.parent.document.body); } catch(e) {}
        try { applyFont(window.top.document.body);    } catch(e) {}
    }
    run();
    try {
        new MutationObserver(run).observe(
            window.parent.document.body, { childList: true, subtree: true }
        );
    } catch(e) {}
})();
</script>
""", unsafe_allow_html=True)

# Fix "uploadupload" icon text in file uploaders
st.markdown("""
<script>
(function fixUploaders() {
    function hide() {
        document.querySelectorAll('[data-testid="stFileUploaderDropzone"] button').forEach(btn => {
            btn.querySelectorAll('[data-testid="stIconMaterial"]').forEach(el => {
                el.style.cssText = 'display:none!important;width:0!important;height:0!important;overflow:hidden!important;font-size:0!important;';
            });
            btn.childNodes.forEach(node => {
                if (node.nodeType === Node.TEXT_NODE && node.textContent.trim().toLowerCase() === 'upload') {
                    node.textContent = '';
                }
            });
        });
    }
    hide();
    new MutationObserver(hide).observe(document.body, { childList: true, subtree: true });
})();
</script>
""", unsafe_allow_html=True)

# ── Fixed Header ──────────────────────────────────────────────────────────────
logo_path = str(LOGO_PATH)
logo_b64  = img_to_base64(logo_path) if Path(logo_path).exists() else ""

st.markdown(
    f"""
    <div class="ark-nav">
      <div class="ark-nav-inner">
        <div class="ark-nav-left">
          <img src="data:image/png;base64,{logo_b64}"
            style="height:68px; width:auto; display:block;" />
          <div>
            <div class="ark-nav-title">M&amp;V Report Sanity Check</div>
            <div class="ark-nav-subtitle">Report vs Calculation Sheet Cross-Reference</div>
          </div>
        </div>
        <div class="ark-nav-right">
          <div class="pill">AI Assistant</div>
        </div>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ── Upload Section ────────────────────────────────────────────────────────────
st.markdown(
    """
    <div class="ark-section"><div class="ark-section-title">Upload documents</div></div>
    <div class="ark-section-rule"></div>
    """,
    unsafe_allow_html=True,
)

col1, col2 = st.columns(2)
with col1:
    report_upload = st.file_uploader(
        "M&V Report (PDF)",
        type=["pdf"],
        accept_multiple_files=False,
        help="The M&V Report PDF submitted by the ESP.",
    )
with col2:
    calc_upload = st.file_uploader(
        "M&V Calculation Sheet (Excel)",
        type=["xlsx"],
        accept_multiple_files=False,
        help="The M&V Calculation Sheet (.xlsx) submitted by the ESP.",
    )

components.html("""
<script>
(function fixUploadersFromComponent() {
    function hide(root) {
        root.querySelectorAll('[data-testid="stIconMaterial"]').forEach(el => {
            el.style.cssText = 'display:none!important;width:0!important;overflow:hidden!important;font-size:0!important;';
        });
        root.querySelectorAll('[data-testid="stFileUploaderDropzone"] button').forEach(btn => {
            btn.childNodes.forEach(node => {
                if (node.nodeType === 3 && node.textContent.trim().toLowerCase() === 'upload') {
                    node.textContent = '';
                }
            });
        });
    }
    function run() {
        try { hide(window.parent.document.body); } catch(e) {}
        try { hide(window.top.document.body); } catch(e) {}
    }
    run();
    try {
        new MutationObserver(run).observe(window.parent.document.body, { childList:true, subtree:true });
    } catch(e) {}
})();
</script>
""", height=0)

# ── Auto-extract submission details when both files are uploaded ───────────────
# @st.cache_data caches by (calc_text, pdf_bytes) — same files = instant cache hit,
# different files = fresh extraction.  The generation counter (_field_gen) changes
# the widget key so Streamlit treats the inputs as brand-new and honours value=.

_prefill = {"client_name": "", "facility_name": "", "country": "", "esp_name": ""}

if report_upload is not None and calc_upload is not None:
    _file_key = f"{report_upload.name}_{report_upload.size}|{calc_upload.name}_{calc_upload.size}"

    if st.session_state.get("_meta_file_key") != _file_key:
        with st.spinner("Extracting submission details from documents…"):
            _pdf_data  = report_upload.read()
            _xlsx_data = calc_upload.read()
            report_upload.seek(0)
            calc_upload.seek(0)
            _calc_text = _xlsx_to_text(_xlsx_data)
            # Step 1: fast direct scan of Excel header rows
            _xlsx_scan = _scan_xlsx_for_metadata(_xlsx_data)
            _prefill = dict(_xlsx_scan)
            # Step 2: Haiku fills remaining gaps
            _haiku = _extract_metadata(_calc_text, _pdf_data)
            for _k, _v in _haiku.items():
                if _v and not _prefill.get(_k):
                    _prefill[_k] = _v

        # Map period_type "M"/"Q" → "Month"/"Quarter" for the UI dropdown
        _pt_raw = _prefill.get("period_type", "") or ""
        _pt_ui  = "Month" if _pt_raw == "M" else ("Quarter" if _pt_raw == "Q" else "")

        # Normalise mv_option to match the selectbox options list
        _mv_raw = (_prefill.get("mv_option", "") or "").strip()
        _MV_NORM = {
            "option a": "Option A", "option b": "Option B",
            "option c mean model": "Option C Mean Model",
            "option c - mean model": "Option C Mean Model",
            "option c": "Option C", "option d": "Option D",
        }
        _mv_ui = _MV_NORM.get(_mv_raw.lower(), _mv_raw if _mv_raw else "")

        st.session_state["_meta_file_key"]       = _file_key
        st.session_state["_prefill_client"]      = _prefill.get("client_name",   "")
        st.session_state["_prefill_facility"]    = _prefill.get("facility_name", "")
        st.session_state["_prefill_country"]     = _prefill.get("country",       "")
        st.session_state["_prefill_esp"]         = _prefill.get("esp_name",      "")
        st.session_state["_prefill_mv_option"]   = _mv_ui
        st.session_state["_prefill_year_number"] = str(_prefill.get("year_number",   "") or "")
        st.session_state["_prefill_period_type"] = _pt_ui
        st.session_state["_prefill_period_num"]  = str(_prefill.get("period_number", "") or "")
        # Bump generation → widget keys change → value= fires as first-render
        st.session_state["_field_gen"] = st.session_state.get("_field_gen", 0) + 1
        st.rerun()

    else:
        # Same files already extracted — read cached prefills from session state
        _prefill = {
            "client_name":   st.session_state.get("_prefill_client",   ""),
            "facility_name": st.session_state.get("_prefill_facility", ""),
            "country":       st.session_state.get("_prefill_country",  ""),
            "esp_name":      st.session_state.get("_prefill_esp",      ""),
        }
else:
    # One or both files removed — clear state and reset fields
    if st.session_state.get("_meta_file_key"):
        for _k in ("_meta_file_key", "_prefill_client", "_prefill_facility",
                   "_prefill_country", "_prefill_esp", "_prefill_mv_option",
                   "_prefill_year_number", "_prefill_period_type", "_prefill_period_num"):
            st.session_state.pop(_k, None)
        st.session_state["_field_gen"] = st.session_state.get("_field_gen", 0) + 1
        st.rerun()

_gen = st.session_state.get("_field_gen", 0)

# ── Submission Details ────────────────────────────────────────────────────────
st.markdown(
    """
    <div class="ark-section"><div class="ark-section-title">Submission details</div></div>
    <div class="ark-section-rule"></div>
    """,
    unsafe_allow_html=True,
)

_sd_c1, _sd_c2, _sd_c3, _sd_c4 = st.columns(4)
with _sd_c1:
    client_name   = st.text_input("Client Name",   key=f"sd_client_{_gen}",   value=st.session_state.get("_prefill_client",   ""))
with _sd_c2:
    facility_name = st.text_input("Facility Name", key=f"sd_facility_{_gen}", value=st.session_state.get("_prefill_facility", ""))
with _sd_c3:
    country       = st.text_input("Country",       key=f"sd_country_{_gen}",  value=st.session_state.get("_prefill_country",  ""))
with _sd_c4:
    esp_name      = st.text_input("ESP's Name",    key=f"sd_esp_{_gen}",      value=st.session_state.get("_prefill_esp",      ""))

# ── Run button ────────────────────────────────────────────────────────────────
btn_left, btn_right = st.columns([7, 3])
with btn_right:
    run_btn = st.button(
        "Generate Comments",
        type="primary",
        disabled=(report_upload is None or calc_upload is None),
        use_container_width=True,
    )

# ── Validation & run ──────────────────────────────────────────────────────────
if run_btn:
    if report_upload is None:
        st.error("M&V Report PDF is required.")
        st.stop()
    if calc_upload is None:
        st.error("M&V Calculation Sheet (Excel) is required.")
        st.stop()

    log_box = st.empty()

    try:
        pdf_bytes  = report_upload.read()
        xlsx_bytes = calc_upload.read()

        log_box.info(
            f"Cross-referencing **{report_upload.name}** against "
            f"**{calc_upload.name}** "
        )

        with st.spinner(
            "Running sanity check — cross-referencing all key values between both documents… "
        ):
            result = run_sanity_check(
                pdf_bytes, xlsx_bytes, report_upload.name,
                xlsx_filename=calc_upload.name,
                submission_details={
                    "client_name":   client_name.strip(),
                    "facility_name": facility_name.strip(),
                    "country":       country.strip(),
                    "esp_name":      esp_name.strip(),
                },
            )

        log_box.empty()

        st.markdown(
            """
            <div class="ark-section"><div class="ark-section-title">Results</div></div>
            <div class="ark-section-rule"></div>
            """,
            unsafe_allow_html=True,
        )
        st.success("Sanity check completed successfully.")

        c1, c2, c3, c4 = st.columns(4)
        c1.markdown(f'<div class="stat-card"><div class="stat-number color-blue">{result["total"]}</div><div class="stat-label">Questions Reviewed</div></div>', unsafe_allow_html=True)
        c2.markdown(f'<div class="stat-card"><div class="stat-number color-green">{result["approved"]}</div><div class="stat-label">Approved</div></div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="stat-card"><div class="stat-number color-red">{result["not_approved"]}</div><div class="stat-label">Not Approved</div></div>', unsafe_allow_html=True)
        c4.markdown(f'<div class="stat-card"><div class="stat-number color-orange">{result["incomplete"]}</div><div class="stat-label">Incomplete</div></div>', unsafe_allow_html=True)

        if result["missing_sns"]:
            st.warning(
                f"**Warning — missing SNs:** The following questions were not returned by the AI "
                f"and have been left blank in the output: **{', '.join(result['missing_sns'])}**"
            )

        st.download_button(
            "⬇️ Download Filled Review Sheet",
            data=result["excel_bytes"],
            file_name=result["filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        log_box.empty()
        st.error(f"{type(e).__name__}: {e}")
