"""
excel_writer.py
---------------
All openpyxl logic for writing AI review results into the M&V Review Sheet Template.
The AI never sees or touches this file — Python handles everything.

Sheet  : "1. M&V Report Compliance Check"
Columns written (1-based):
  H  col 8  = Active Status
  I  col 9  = Consultant Comments

Colour coding:
  Approved     -> bg #C6EFCE  font #375623  bold  centered
  Not Approved -> bg #FFC7CE  font #9C0006  bold  centered
  Incomplete   -> bg #FFEB9C  font #9C6500  bold  centered

  Comment non-empty -> bg #FFFF99  font #000000  wrap  top-aligned
  Comment empty     -> bg #FFFFFF
  Row height        -> max(50, (len//80 + newlines + 1) * 15)

  All written cells: Calibri 11, thin border #BFBFBF.
"""

import io
import zipfile
import re
from datetime import date

import openpyxl
import openpyxl.formatting.formatting
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Column indices (1-based) ─────────────────────────────────────────────────
COL_SN      = 2   # B
COL_STATUS  = 8   # H  Active Status
COL_COMMENT = 9   # I  Consultant Comments
DATA_START  = 24  # First data row to scan

FONT_NAME    = "Trebuchet MS"
FONT_SIZE    = 11
BORDER_COLOR = "BFBFBF"

# ── Style maps ───────────────────────────────────────────────────────────────
STATUS_STYLES = {
    "Approved":     {"bg": "C6EFCE", "fc": "375623"},
    "Not Approved": {"bg": "FFC7CE", "fc": "9C0006"},
    "Incomplete":   {"bg": "FFEB9C", "fc": "9C6500"},
}

COMMENT_BG = "FFFF99"
COMMENT_FC = "000000"


def _make_border() -> Border:
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _style_cell(cell, value: str, bg: str, fc: str,
                bold: bool = False, wrap: bool = False,
                align: str = "left") -> None:
    cell.value = value
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE, color=fc, bold=bold)
    cell.fill = _make_fill(bg)
    vertical = "center" if align == "center" else "top"
    cell.alignment = Alignment(horizontal=align, vertical=vertical, wrap_text=wrap)
    cell.border = _make_border()


def _is_section_header(value) -> bool:
    if value is None:
        return True
    s = str(value).strip()
    if not s:
        return True
    try:
        f = float(s)
        return f == int(f) and "." not in s
    except (ValueError, TypeError):
        return False


def _fill_cover_page(cp, meta: dict) -> None:
    """
    Write extracted metadata into the Cover Page sheet.

    Layout (1-based):
      A10  (R10C1) : "[Client_Name, Facility, Country]"
      A11  (R11C1) : "Energy Efficiency Retrofit Project for [project_name]"
      B13  (R13C2) : Year number → "Y#"
      D13  (R13C4) : Month / Quarter → "M#" or "Q#"
      B14  (R14C2) : Date of review (today dd/mm/yy)
      B15  (R15C2) : ESP name
      C16  (R16C3) : M&V Option
      F19  (R19C6) : Date of Last Status (today)
      B25  (R25C2) : Consultant sign-off date (today)
    """
    today_str = date.today().strftime("%d/%m/%y")

    # A10 — client name, facility name, country
    parts = [
        meta.get("client_name", ""),
        meta.get("facility_name", ""),
        meta.get("country", ""),
    ]
    a10 = ", ".join(p for p in parts if p)
    if a10:
        cp.cell(row=10, column=1).value = a10

    # A11 — project title
    project_name = meta.get("project_name", "")
    if project_name:
        cp.cell(row=11, column=1).value = (
            f"Energy Efficiency Retrofit Project for {project_name}"
        )

    # B13 — year number as "Y#"
    year_num = meta.get("year_number")
    if year_num is not None:
        try:
            cp.cell(row=13, column=2).value = f"Y{int(year_num)}"
        except (ValueError, TypeError):
            pass

    # D13 — month or quarter as "M#" / "Q#"
    period_type   = meta.get("period_type", "")
    period_number = meta.get("period_number")
    if period_type and period_number is not None:
        try:
            cp.cell(row=13, column=4).value = f"{period_type.upper()}{int(period_number)}"
        except (ValueError, TypeError):
            pass

    # B14 — date of review
    cp.cell(row=14, column=2).value = today_str

    # B15 — ESP name
    esp_name = meta.get("esp_name", "")
    if esp_name:
        cp.cell(row=15, column=2).value = esp_name

    # C16 — M&V option
    mv_option = meta.get("mv_option", "")
    if mv_option:
        cp.cell(row=16, column=3).value = mv_option

    # F19 — Date of Last Status
    cp.cell(row=19, column=6).value = today_str

    # B25 — Consultant sign-off date
    cp.cell(row=25, column=2).value = f"Date: {today_str}"


def _overall_assessment(review_by_sn: dict) -> str:
    """Derive overall assessment from the review results."""
    statuses = [item.get("status", "") for item in review_by_sn.values()]
    if any(s == "Not Approved" for s in statuses):
        return "Not Approved"
    if any(s == "Incomplete" for s in statuses):
        return "Incomplete"
    if statuses and all(s == "Approved" for s in statuses):
        return "Approved"
    return "Incomplete"


def _fill_compliance_sheet_header(ws, meta: dict, review_by_sn: dict) -> None:
    """
    Fill the header section of the '1. M&V Report Compliance Check' sheet.

    Layout (col letters → 1-based col numbers):
      C3  (R3  C3) : Client Name, Facility, Country
      C4  (R4  C3) : Energy Efficiency Retrofit Project for [project_name]
      E2  (R2  C5) : M&V Option
      C6  (R6  C3) : Updated: dd/mm/yy
      D9  (R9  C4) : M&V Reporting period (Start date)
      D10 (R10 C4) : M&V Reporting period (End date)
      B14 (R14 C2) : Round number (1)
      C14 (R14 C3) : Issued On (today)
      D14 (R14 C4) : Received on (today)
      E14 (R14 C5) : Reviewed on (today)
      F14 (R14 C6) : Assessment (derived from results)
    """
    today_str = date.today().strftime("%d/%m/%y")

    # C3 — client name, facility, country
    parts = [meta.get("client_name", ""), meta.get("facility_name", ""), meta.get("country", "")]
    a3 = ", ".join(p for p in parts if p)
    if a3:
        ws.cell(row=3, column=3).value = a3

    # C4 — project title
    project_name = meta.get("project_name", "")
    if project_name:
        ws.cell(row=4, column=3).value = f"Energy Efficiency Retrofit Project for {project_name}"

    # E2 — M&V option (the "Select from List" cell)
    mv_option = meta.get("mv_option", "")
    if mv_option:
        ws.cell(row=2, column=5).value = mv_option

    # C6 — Updated date
    ws.cell(row=6, column=3).value = f" Updated: {today_str}"

    # D9 — reporting period start
    start = meta.get("reporting_period_start", "")
    if start:
        ws.cell(row=9, column=4).value = start

    # D10 — reporting period end
    end = meta.get("reporting_period_end", "")
    if end:
        ws.cell(row=10, column=4).value = end

    # Round 1 row (row 14) — Round, Issued On, Received on, Reviewed on, Assessment
    ws.cell(row=14, column=2).value = 1
    ws.cell(row=14, column=3).value = today_str
    ws.cell(row=14, column=4).value = today_str
    ws.cell(row=14, column=5).value = today_str

    assessment = _overall_assessment(review_by_sn)
    assess_s = STATUS_STYLES.get(assessment, {"bg": "FFFFFF", "fc": "000000"})
    _style_cell(
        ws.cell(row=14, column=6),
        assessment, bg=assess_s["bg"], fc=assess_s["fc"],
        bold=True, align="center"
    )


def write_review(template_bytes: bytes, review_by_sn: dict, meta: dict = None) -> bytes:
    """
    Load the Excel template from bytes, write review results, return as bytes.

    Parameters
    ----------
    template_bytes : bytes
        Raw bytes of the M&V Review Sheet Template.
    review_by_sn  : dict
        Dict keyed by SN string -> {"status": ..., "comment": ...}
    meta : dict, optional
        Extracted metadata for the cover page.

    Returns
    -------
    bytes
        In-memory bytes of the filled workbook (xlsx).
    """
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes), keep_links=False, keep_vba=False)

    # ── Aggressive cleanup to prevent Excel repair warnings ───────────────────
    # 1. External links
    wb._external_links.clear()

    # 2. All defined names / named ranges — clear in-place (never replace the
    #    object; openpyxl's serialiser holds a reference and breaks otherwise).
    #    The XML post-processor strips <definedNames> as a hard backstop.
    for _attr in ("_list", "_dict", "definedName"):
        try:
            _container = getattr(wb.defined_names, _attr, None)
            if _container is not None:
                _container.clear()
                break
        except Exception:
            pass

    for sheet in wb.worksheets:
        # 3. Charts and images
        sheet._charts.clear()
        sheet._images.clear()
        if hasattr(sheet, "_drawing") and sheet._drawing is not None:
            sheet._drawing = None

        # 4. Data validations (broken formula references cause repair warnings)
        try:
            sheet.data_validations.dataValidation.clear()
        except Exception:
            pass

        # 5. Conditional formatting (external/broken refs cause warnings)
        try:
            sheet.conditional_formatting._cf_rules.clear()
        except Exception:
            try:
                sheet.conditional_formatting = openpyxl.formatting.formatting.ConditionalFormattingList()
            except Exception:
                pass

        # 6. Pivot tables
        try:
            sheet._pivots.clear()
        except Exception:
            pass

        # 7. Table objects (ListObjects) — broken table refs are a common cause
        try:
            for tbl_name in list(sheet.tables):
                del sheet.tables[tbl_name]
        except Exception:
            try:
                sheet._tables.clear()
            except Exception:
                pass

        # 8. Sparklines
        try:
            sheet._sparkline_groups.clear()
        except Exception:
            pass

        # 9. Legacy drawings (VML — comments, form controls)
        try:
            sheet.legacy_drawing = None
        except Exception:
            pass

        # 10. Print areas and titles that may reference broken ranges
        try:
            sheet.print_area = None
        except Exception:
            pass

    # ── Cover Page ────────────────────────────────────────────────────────────
    if meta and "Cover Page" in wb.sheetnames:
        _fill_cover_page(wb["Cover Page"], meta)

    ws = wb["1. M&V Report Compliance Check"]

    # ── Compliance sheet header ───────────────────────────────────────────────
    if meta:
        _fill_compliance_sheet_header(ws, meta, review_by_sn)

    for row_idx in range(DATA_START, ws.max_row + 1):
        sn_val = ws.cell(row=row_idx, column=COL_SN).value
        if _is_section_header(sn_val):
            continue

        sn = str(sn_val).strip()
        if sn not in review_by_sn:
            continue

        item    = review_by_sn[sn]
        status  = item.get("status", "")
        comment = item.get("comment", "") or ""

        # ── Active Status (col H) ────────────────────────────────────────────
        st_s = STATUS_STYLES.get(status, {"bg": "FFFFFF", "fc": "000000"})
        _style_cell(
            ws.cell(row=row_idx, column=COL_STATUS),
            status, bg=st_s["bg"], fc=st_s["fc"],
            bold=True, align="center"
        )

        # ── Consultant Comments (col I) ──────────────────────────────────────
        if comment:
            _style_cell(
                ws.cell(row=row_idx, column=COL_COMMENT),
                comment, bg=COMMENT_BG, fc=COMMENT_FC,
                wrap=True, align="left"
            )
            lines = len(comment) // 80 + comment.count("\n") + 1
            ws.row_dimensions[row_idx].height = max(50, lines * 15)
        else:
            _style_cell(
                ws.cell(row=row_idx, column=COL_COMMENT),
                "", bg="FFFFFF", fc=COMMENT_FC,
                wrap=True, align="left"
            )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    raw = out.read()
    return _strip_xlsx_junk(raw)


# ── Parts that reliably cause Excel repair warnings when openpyxl writes them ─
_STRIP_PARTS = {
    "xl/calcChain.xml",         # stale formula calculation cache → #1 cause
    "xl/sharedStrings.xml",     # can be regenerated; sometimes has corrupt refs
}
# Regex to remove a <Relationship .../> entry referencing a stripped part
_REL_RE = re.compile(
    r'<Relationship\s[^>]*Target="[^"]*(?:calcChain|sharedStrings)[^"]*"[^/]*/?>',
    re.IGNORECASE,
)
# Regex to remove a <Override .../> entry in [Content_Types].xml
_CT_RE = re.compile(
    r'<Override\s[^>]*PartName="[^"]*(?:calcChain|sharedStrings)[^"]*"[^/]*/?>',
    re.IGNORECASE,
)
# Strip the entire <definedNames>…</definedNames> block from workbook.xml
# (broken named ranges are the exact cause of this repair warning)
_DEFINED_NAMES_RE = re.compile(
    r'<definedNames\b[^>]*>.*?</definedNames>|<definedNames\b[^/]*/?>',
    re.IGNORECASE | re.DOTALL,
)


def _strip_xlsx_junk(xlsx_bytes: bytes) -> bytes:
    """
    Post-process the saved workbook zip:
    - Remove xl/calcChain.xml  (stale calc chain → Excel repair warning)
    - Remove xl/sharedStrings.xml  (openpyxl sometimes writes a bad one)
    - Strip <definedNames> block from xl/workbook.xml (broken named ranges)
    - Patch xl/_rels/workbook.xml.rels and [Content_Types].xml to remove
      references to deleted files so Excel doesn't look for them.
    """
    src = io.BytesIO(xlsx_bytes)
    dst = io.BytesIO()

    with zipfile.ZipFile(src, "r") as zin, \
         zipfile.ZipFile(dst, "w", compression=zipfile.ZIP_DEFLATED) as zout:

        for item in zin.infolist():
            name = item.filename

            # Skip the parts we're stripping
            if name in _STRIP_PARTS:
                continue

            data = zin.read(name)

            # Strip definedNames block from workbook.xml
            if name == "xl/workbook.xml":
                text = data.decode("utf-8", errors="replace")
                text = _DEFINED_NAMES_RE.sub("", text)
                data = text.encode("utf-8")

            # Patch the workbook relationships file
            if name == "xl/_rels/workbook.xml.rels":
                text = data.decode("utf-8", errors="replace")
                text = _REL_RE.sub("", text)
                data = text.encode("utf-8")

            # Patch [Content_Types].xml
            if name == "[Content_Types].xml":
                text = data.decode("utf-8", errors="replace")
                text = _CT_RE.sub("", text)
                data = text.encode("utf-8")

            zout.writestr(item, data)

    dst.seek(0)
    return dst.read()
